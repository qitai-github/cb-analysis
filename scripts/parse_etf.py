#!/usr/bin/env python3
"""
解析 ETF 持股 xlsx 檔案，統一輸出為 data/etf-holdings.json
支援 7 檔 ETF 的不同格式，自動偵測欄位位置

用法：python scripts/parse_etf.py
"""

import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("需要安裝 openpyxl: pip install openpyxl")
    sys.exit(1)

ETF_DIR = Path(__file__).parent.parent / "ETF"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "etf-holdings.json"

# ETF 設定：代碼 -> 檔名 pattern, 名稱, 解析器類型
ETF_CONFIG = {
    "00891": {"pattern": "00891", "name": "中信關鍵半導體", "parser": "ctbc"},
    "00881": {"pattern": "CR", "name": "國泰科技龍頭", "parser": "cathay"},
    "00981A": {"pattern": "00981A", "name": "統一台股增長", "parser": "ezmoney"},
    "00988A": {"pattern": "00988A", "name": "統一全球創新", "parser": "ezmoney"},
    "00991A": {"pattern": "復華", "name": "復華未來50", "parser": "fhtrust"},
    "00982A": {"pattern": "00982A", "name": "群益台灣強棒", "parser": "capital"},
    "00992A": {"pattern": "00992A", "name": "群益科技創新", "parser": "capital"},
}


def find_etf_file(pattern):
    """找到 ETF 目錄中匹配的 xlsx 檔案"""
    for f in ETF_DIR.glob("*.xlsx"):
        if pattern in f.name:
            return f
    return None


def clean_number(val):
    """清理數字字串，移除逗號和貨幣符號"""
    if val is None:
        return None
    s = str(val).replace(",", "").replace("NTD", "").replace("NT$", "").replace("TWD", "").replace("USD", "").strip()
    s = re.sub(r"[$ ]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def extract_date_from_sheet(ws, parser_type):
    """從工作表中提取資料日期"""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            # 西元格式優先: 2026/04/17 or 2026-04-17
            m = re.search(r"(20\d{2})[/-](\d{2})[/-](\d{2})", s)
            if m:
                return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            # 格式: 2026_04_17
            m = re.search(r"(20\d{2})_(\d{2})_(\d{2})", s)
            if m:
                return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            # 民國年格式: 115/04/17 (前面不能有數字，避免匹配 2026 的 026)
            m = re.search(r"(?<!\d)(\d{2,3})/(\d{2})/(\d{2})", s)
            if m and int(m.group(1)) >= 100 and int(m.group(1)) <= 200:
                y = int(m.group(1)) + 1911
                return f"{y}-{m.group(2)}-{m.group(3)}"
    return None


def extract_nav_info(ws):
    """提取基金淨資產資訊"""
    nav = None
    units = None
    nav_per_unit = None
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        for i, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell)
            if "淨資產" in s or "資產淨值" in s:
                # 值可能在同行下一格或本格含數字
                val = row[i + 1] if i + 1 < len(row) else None
                if val:
                    nav = clean_number(val)
            if "流通" in s and ("單位" in s or "總數" in s):
                val = row[i + 1] if i + 1 < len(row) else None
                if val:
                    units = clean_number(val)
            if "每" in s and "淨值" in s:
                val = row[i + 1] if i + 1 < len(row) else None
                if val:
                    nav_per_unit = clean_number(val)
    return {"nav": nav, "units": units, "navPerUnit": nav_per_unit}


def parse_ctbc(filepath):
    """中信 00891 格式：支援兩種版本
    舊版：序號/代碼/中文名稱/英文名稱/股數/權重(%)
    新版：股票代號/股票名稱/股數/持股權重
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    date = extract_date_from_sheet(ws, "ctbc")
    nav_info = extract_nav_info(ws)
    holdings = []

    found_header = False
    code_col = 1  # default: 舊版格式 code 在 col 1
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not found_header:
            r0 = str(row[0]).strip() if row[0] else ""
            if r0 == "序號":
                found_header = True
                code_col = 1
            elif r0 == "股票代號":
                found_header = True
                code_col = 0
            continue
        if row[0] is None or str(row[0]).strip() == "":
            continue

        if code_col == 0:
            # 新版：股票代號(0), 股票名稱(1), 股數(2), 持股權重(3)
            code = str(row[0]).strip()
            if not code or not code[0].isdigit():
                continue
            name = str(row[1]).strip() if row[1] else ""
            shares = clean_number(row[2])
            weight_str = str(row[3]).replace("%", "").strip() if row[3] else ""
            weight = clean_number(weight_str)
        else:
            # 舊版：序號(0), 代碼(1), 名稱(2), ?(3), 股數(4), 權重(5)
            code = str(row[1]).strip() if row[1] else ""
            if not code or not code[0].isdigit():
                continue  # 跳過期貨(TX/TE)、現金(TWD)、表頭等非股票項目
            name = str(row[2]).strip() if row[2] else ""
            shares = clean_number(row[4]) if len(row) > 4 else None
            weight = clean_number(row[5]) if len(row) > 5 else None

        if code:
            holdings.append({"code": code, "name": name, "shares": shares, "weight": weight})

    return {"date": date, **nav_info, "holdings": holdings}


def parse_cathay(filepath):
    """國泰 00881 格式：代號/名稱/股數/持股權重 (股票區塊在「股票」標題行之後)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    date = extract_date_from_sheet(ws, "cathay")
    nav_info = extract_nav_info(ws)
    holdings = []

    found_header = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        # header 行: 第一格剛好是「股票代號」（不是包含在長字串中）
        if row[0] and str(row[0]).strip() == "股票代號":
            found_header = True
            continue
        if not found_header:
            continue
        code = str(row[0]).strip() if row[0] else ""
        if not code or "期貨" in code:
            break
        name = str(row[1]).strip() if row[1] else ""
        shares = clean_number(row[2])
        weight_str = str(row[3]).replace("%", "").strip() if row[3] else ""
        weight = clean_number(weight_str)
        if code:
            holdings.append({"code": code, "name": name, "shares": shares, "weight": weight})

    return {"date": date, **nav_info, "holdings": holdings}


def parse_ezmoney(filepath):
    """統一 00981A/00988A 格式：代號/名稱/股數/持股權重"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    date = extract_date_from_sheet(ws, "ezmoney")
    nav_info = extract_nav_info(ws)
    holdings = []

    found_header = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and "股票代號" in str(row[0]):
            found_header = True
            continue
        if not found_header:
            continue
        code = str(row[0]).strip() if row[0] else ""
        if not code:
            break
        name = str(row[1]).strip() if row[1] else ""
        shares = clean_number(row[2])
        weight_str = str(row[3]).replace("%", "").strip() if row[3] else ""
        weight = clean_number(weight_str)
        if code:
            holdings.append({"code": code, "name": name, "shares": shares, "weight": weight})

    return {"date": date, **nav_info, "holdings": holdings}


def parse_fhtrust(filepath):
    """復華 00991A 格式：代碼/名稱/股數/金額/權重(%)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    date = extract_date_from_sheet(ws, "fhtrust")

    # 復華的 NAV 資訊格式不同，手動解析
    nav = None
    units = None
    nav_per_unit = None
    rows_data = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    for i, row in enumerate(rows_data):
        if row[0] and "資產淨值" in str(row[0]):
            if i + 1 < len(rows_data):
                nav = clean_number(rows_data[i + 1][0])
        if row[0] and "流通單位" in str(row[0]):
            if i + 1 < len(rows_data):
                units = clean_number(rows_data[i + 1][0])
        if row[0] and "每單位淨值" in str(row[0]):
            if i + 1 < len(rows_data):
                nav_per_unit = clean_number(rows_data[i + 1][0])

    holdings = []
    found_header = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and "證券代號" in str(row[0]):
            found_header = True
            continue
        if not found_header:
            continue
        code = str(row[0]).strip() if row[0] else ""
        if not code:
            break
        name = str(row[1]).strip() if row[1] else ""
        shares = clean_number(row[2])
        weight_str = str(row[4]).replace("%", "").strip() if row[4] else ""
        weight = clean_number(weight_str)
        if code:
            holdings.append({"code": code, "name": name, "shares": shares, "weight": weight})

    return {"date": date, "nav": nav, "units": units, "navPerUnit": nav_per_unit, "holdings": holdings}


def parse_capital(filepath):
    """群益 00982A/00992A 格式：多 sheet，股票在「股票」工作表"""
    wb = openpyxl.load_workbook(filepath)

    # 從「投資組合」sheet 取 NAV 資訊
    ws_main = wb["投資組合"]
    nav = None
    units = None
    nav_per_unit = None
    for row in ws_main.iter_rows(min_row=1, values_only=True):
        label = str(row[0]) if row[0] else ""
        val = row[1] if len(row) > 1 else None
        if "基金淨資產" in label and "每" not in label:
            nav = clean_number(val)
        elif "每受益權" in label:
            nav_per_unit = clean_number(val)
        elif "總數" in label:
            units = clean_number(val)

    # 從檔名抓日期 (群益 xlsx 內沒有日期欄位)
    date = None
    fname = filepath.name
    m = re.search(r"(20\d{2})[_-]?(\d{2})[_-]?(\d{2})", fname)
    if m:
        date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    # 從「股票」sheet 取持股
    ws_stock = wb["股票"]
    holdings = []
    found_header = False
    for row in ws_stock.iter_rows(min_row=1, values_only=True):
        if row[0] and "股票代號" in str(row[0]):
            found_header = True
            continue
        if not found_header:
            continue
        code = str(row[0]).strip() if row[0] else ""
        if not code:
            break
        name = str(row[1]).strip() if row[1] else ""
        weight_str = str(row[2]).replace("%", "").strip() if row[2] else ""
        weight = clean_number(weight_str)
        shares = clean_number(row[3])
        if code:
            holdings.append({"code": code, "name": name, "shares": shares, "weight": weight})

    return {"date": date, "nav": nav, "units": units, "navPerUnit": nav_per_unit, "holdings": holdings}


PARSERS = {
    "ctbc": parse_ctbc,
    "cathay": parse_cathay,
    "ezmoney": parse_ezmoney,
    "fhtrust": parse_fhtrust,
    "capital": parse_capital,
}


def compute_holding_changes(etfs, prev_etfs):
    """比對每檔 ETF 的持股變動"""
    for etf_code, etf in etfs.items():
        prev = prev_etfs.get(etf_code) if prev_etfs else None
        prev_map = {}
        if prev and prev.get("holdings"):
            for h in prev["holdings"]:
                prev_map[h["code"]] = h

        summary = {"added": 0, "removed": 0, "increased": 0, "decreased": 0, "unchanged": 0}
        etf["prevDate"] = prev["date"] if prev else None

        for h in etf["holdings"]:
            ph = prev_map.pop(h["code"], None)
            if not ph:
                h["change"] = "added"
                h["prevWeight"] = None
                h["prevShares"] = None
                summary["added"] += 1
            else:
                h["prevWeight"] = ph.get("weight")
                h["prevShares"] = ph.get("shares")
                if h.get("weight") is not None and ph.get("weight") is not None:
                    diff = abs(h["weight"] - ph["weight"])
                    if diff < 0.005:
                        h["change"] = "unchanged"
                        summary["unchanged"] += 1
                    elif h["weight"] > ph["weight"]:
                        h["change"] = "increased"
                        summary["increased"] += 1
                    else:
                        h["change"] = "decreased"
                        summary["decreased"] += 1
                else:
                    h["change"] = "unchanged"
                    summary["unchanged"] += 1

        # 被刪除的持股（過濾非股票代碼如期貨TX/TE、現金TWD等）
        removed = []
        for rcode, rh in prev_map.items():
            if not rcode or not rcode[0].isdigit():
                continue
            removed.append({
                "code": rh["code"],
                "name": rh.get("name", ""),
                "shares": None,
                "weight": None,
                "change": "removed",
                "prevWeight": rh.get("weight"),
                "prevShares": rh.get("shares"),
            })
            summary["removed"] += 1

        etf["holdings"].extend(removed)
        etf["holdingCount"] = len(etf["holdings"]) - len(removed)
        etf["changes"] = summary

        print(f"  {etf_code} 變動: +{summary['added']} -{summary['removed']} ↑{summary['increased']} ↓{summary['decreased']} ={summary['unchanged']}")


def main():
    etfs = {}

    for etf_code, config in ETF_CONFIG.items():
        filepath = find_etf_file(config["pattern"])
        if not filepath:
            print(f"[跳過] {etf_code} {config['name']}：找不到檔案 (pattern: {config['pattern']})")
            continue

        print(f"[解析] {etf_code} {config['name']}：{filepath.name}")
        parser = PARSERS[config["parser"]]
        try:
            result = parser(filepath)
        except Exception as e:
            print(f"  [錯誤] {e}")
            continue

        etf_info = {
            "name": config["name"],
            "date": result["date"],
            "nav": result.get("nav"),
            "units": result.get("units"),
            "navPerUnit": result.get("navPerUnit"),
            "holdingCount": len(result["holdings"]),
            "holdings": result["holdings"],
        }
        etfs[etf_code] = etf_info
        print(f"  日期: {result['date']}, 持股: {len(result['holdings'])} 檔")

    # 讀取前一版 JSON 進行比對
    prev_etfs = None
    if OUTPUT_PATH.exists():
        try:
            with open(OUTPUT_PATH, "r", encoding="utf-8") as f:
                prev_data = json.load(f)
                prev_etfs = prev_data.get("etfs")
            print(f"\n讀取前一版資料進行比對...")
        except Exception as e:
            print(f"  前一版讀取失敗: {e}")

    compute_holding_changes(etfs, prev_etfs)

    # 合併持股
    merged = {}
    for etf_code, etf_info in etfs.items():
        for h in etf_info["holdings"]:
            if h.get("change") == "removed":
                continue  # 已刪除的不計入 merged
            code = h["code"]
            if code not in merged:
                merged[code] = {"name": h["name"], "etfs": {}, "count": 0}
            merged[code]["etfs"][etf_code] = h.get("weight")
            merged[code]["count"] = len(merged[code]["etfs"])

    # 輸出 JSON
    output = {
        "_meta": {
            "generatedAt": __import__("datetime").datetime.now().isoformat(),
            "etfCount": len(etfs),
            "totalStocks": len(merged),
        },
        "etfs": etfs,
        "merged": merged,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n輸出完成: {OUTPUT_PATH}")
    print(f"共 {len(etfs)} 檔 ETF，{len(merged)} 檔不重複持股")


if __name__ == "__main__":
    main()
