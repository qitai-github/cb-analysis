#!/usr/bin/env python3
"""把人工彙整的「CB重大訊息彙整.xlsx」匯進 data/mops_news.json 的 cbEvents。

為什麼需要這支:`mops_news.py` 的兩個 OpenAPI 只給「最近一個發言日」的快照,
MOPS 逐檔查詢也只回最近數則 → **2026-08-20 之前的公告一律補不回來**。這份
xlsx 是從 MOPS「歷史重大訊息」(114/115 年度) 逐檔查出來的彙整,剛好把那段
歷史補上,而且含董事會決議日 (CBAS 掛牌後就查不到的那一欄,見 FEATURES §8.2.1)。

xlsx 欄位 → 事件型別:
  1.董事會決議日期時間              → board          (事件軸第 1 格)
  2.轉換價格及溢價率公告日期時間    → auctionNotice  (第 3 格)
  3.代收價款行庫及存儲專戶行庫公告  → collection     (第 2 格)

CB 代碼直接取表上的「CB代碼」欄,不用從主旨反推期次 (主旨常一次公告兩檔,
例如「第二次暨第三次無擔保轉換公司債」,反推會錯)。

寫進 cbEvents 後就永久留著:mops_news.py 每次 merge 都會沿用舊的 cbEvents,
再由 parse_and_export Phase 4.75 併進 all-data 的 cbasCalendar.events。

跑法:
  python scripts/import_cb_announcements.py                    # 預設讀專案根目錄那份
  python scripts/import_cb_announcements.py --xlsx 路徑 --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "CB重大訊息彙整.xlsx"
TARGET = ROOT / "data" / "mops_news.json"
TZ8 = timezone(timedelta(hours=8))

# 欄位標題關鍵字 → (事件型別, 主旨欄關鍵字)
COLS = [
    ("board", "董事會決議日期"),
    ("auctionNotice", "轉換價格及溢價率公告日期"),
    ("collection", "代收價款行庫及存儲專戶行庫公告日期"),
]


def log(m: str) -> None:
    print(m, flush=True)


def roc_dt_to_iso(v) -> str:
    """'114/11/07 15:18:23' → '2025-11-07';已經是 datetime 也吃。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    m = re.match(r"^\s*(\d{2,3})/(\d{1,2})/(\d{1,2})", str(v))
    if not m:
        return ""
    y = int(m.group(1)) + 1911
    return f"{y}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"


def parse_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    head = [str(h or "").strip() for h in rows[0]]

    def col(keyword: str) -> int:
        for i, h in enumerate(head):
            if keyword in h:
                return i
        return -1

    i_code, i_name = col("CB代碼"), col("CB簡稱")
    if i_code < 0:
        raise SystemExit(f"找不到「CB代碼」欄,實際表頭:{head}")

    out: list[dict] = []
    for r in rows[1:]:
        if not r or not r[i_code]:
            continue
        cb_code = str(r[i_code]).strip()
        cb_name = str(r[i_name] or "").strip() if i_name >= 0 else ""
        for etype, kw in COLS:
            ci = col(kw)
            if ci < 0 or ci >= len(r):
                continue
            date = roc_dt_to_iso(r[ci])
            if not date:
                continue                      # 空白 / 「(查無資料)」都在這裡濾掉
            title = ""
            if ci + 1 < len(r) and r[ci + 1]:
                title = re.sub(r"\s+", "", str(r[ci + 1]))
            out.append({
                "date": date, "type": etype, "cbCode": cb_code, "cbName": cb_name,
                "stockCode": cb_code[:4], "title": title, "src": "xlsx",
            })
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        raise SystemExit(f"找不到 {path}")
    fresh = parse_xlsx(path)
    by_type: dict[str, int] = {}
    for e in fresh:
        by_type[e["type"]] = by_type.get(e["type"], 0) + 1
    log(f"[1/2] 讀到 {len(fresh):,} 筆事件 {by_type}")

    data = json.loads(TARGET.read_text(encoding="utf-8")) if TARGET.exists() else {}
    data.setdefault("items", [])
    events = data.setdefault("cbEvents", [])
    have = {(e["cbCode"], e["type"], e["date"]) for e in events}
    # 同一 (cbCode, type) 已經有別的日期時要看得出來 — 人工表跟 MOPS 抓的可能撞
    by_key = {}
    for e in events:
        by_key.setdefault((e["cbCode"], e["type"]), e)

    added, dup, conflict = 0, 0, []
    for e in fresh:
        k3 = (e["cbCode"], e["type"], e["date"])
        if k3 in have:
            dup += 1
            continue
        old = by_key.get((e["cbCode"], e["type"]))
        if old:
            conflict.append((e["cbCode"], e["type"], old["date"], e["date"]))
            continue                          # 已經有來源了就不覆蓋,只回報
        events.append(e)
        by_key[(e["cbCode"], e["type"])] = e
        have.add(k3)
        added += 1

    events.sort(key=lambda e: (e.get("date") or "", e.get("cbCode") or ""))
    log(f"[2/2] 新增 {added:,} 筆、重複 {dup} 筆、與既有來源日期不同 {len(conflict)} 筆")
    for c in conflict:
        log(f"      ⚠️  {c[0]} {c[1]}: 既有 {c[2]} / xlsx {c[3]} → 保留既有")

    if args.dry_run:
        log("--dry-run,沒有寫檔")
        return 0
    data["updatedAt"] = datetime.now(TZ8).isoformat(timespec="seconds")
    TARGET.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")),
                      encoding="utf-8")
    log(f"✓ 寫入 {TARGET} (cbEvents={len(events):,})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
