# -*- coding: utf-8 -*-
"""近半年發行 CB → 股號/股名/發行總量(張)/上市櫃後第6個交易日成交量(張)

用法: python scripts/cb_day6_volume.py [起算日 YYYY-MM-DD] [-o 輸出.xlsx]
預設起算日 = 今天往前 6 個月。

資料來源 (data/all-data.json):
  cbasCalendar.issuedInfo  → 發行日(issueDate)、發行總額(actualTotal,百萬元)
  cbDailyTrading           → 每日成交量(張);每檔 5 列一組,只有首列帶代號/名稱
"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "all-data.json"
NTH_DAY = 6  # 上市櫃當日算第 1 天


def load_volume_series(all_data: dict) -> tuple[list[str], dict[str, dict[str, str]]]:
    """回 (交易日列表, {cbCode: {date: 成交量字串}})。"""
    table = all_data["cbDailyTrading"]
    header = table[0]
    dates = header[3:]
    series: dict[str, dict[str, str]] = {}
    for i, row in enumerate(table):
        code = row[0]
        if not code or row[2] != "收盤價":
            continue
        # 同一檔的 5 個類別是連續 5 列,成交量在 +4
        vol_row = table[i + 4]
        if vol_row[2] != "成交量(張)":
            continue
        series[code] = {d: vol_row[3 + j] for j, d in enumerate(dates)}
    return dates, series


def drop_holiday_columns(dates: list[str], series: dict[str, dict[str, str]]) -> tuple[list[str], list[str]]:
    """剔除「全市場零成交」的日期欄。

    這些是舊版假日判斷 bug 留下的殘欄 (例:20260212/20260213 春節連假),
    全欄皆 0。CB 市場每個交易日都有 200 檔上下有成交量,全市場零成交
    不可能發生,故視為非交易日。留著會讓「第 N 個交易日」整個數錯。
    """
    live, dead = [], []
    for d in dates:
        has = any(str(s.get(d, "0")).strip() not in ("", "-", "0") for s in series.values())
        (live if has else dead).append(d)
    return live, dead


def nth_trading_day(dates: list[str], listing: str, n: int) -> str | None:
    """listing (YYYYMMDD) 當日算第 1 天,回第 n 個交易日;不足回 None。"""
    after = [d for d in dates if d >= listing]
    return after[n - 1] if len(after) >= n else None


def main() -> None:
    args = [a for a in sys.argv[1:]]
    out_path = ROOT / "近半年CB發行_上市後第6日成交量.xlsx"
    if "-o" in args:
        i = args.index("-o")
        out_path = Path(args[i + 1])
        del args[i:i + 2]
    if args:
        since = args[0].replace("-", "")
    else:
        t = date.today()
        m, y = t.month - 6, t.year
        if m <= 0:
            m += 12
            y -= 1
        since = f"{y}{m:02d}{t.day:02d}"

    all_data = json.loads(DATA.read_text(encoding="utf-8"))
    issued = all_data["cbasCalendar"]["issuedInfo"]
    dates, series = load_volume_series(all_data)
    dates, dead_dates = drop_holiday_columns(dates, series)

    rows = []
    for code, info in issued.items():
        iso = (info.get("issueDate") or "").replace("-", "")
        if len(iso) != 8 or iso < since:
            continue
        total_lots = info.get("actualTotal")  # 百萬元 → 張 (每張面額 10 萬)
        total_lots = int(round(total_lots * 10)) if total_lots is not None else None
        day6 = nth_trading_day(dates, iso, NTH_DAY)
        vol = note = None
        if code not in series:
            note = "尚未掛牌" if iso >= dates[-1] else "無交易資料"
        elif day6 is None:
            note = f"掛牌未滿 {NTH_DAY} 個交易日"
        else:
            raw = series[code].get(day6, "")
            vol = int(float(raw)) if str(raw).strip() not in ("", "-") else 0
        rows.append({
            "code": code,
            "name": info.get("cbName", ""),
            "total": total_lots,
            "vol": vol,
            "listing": iso,
            "day6": day6,
            "note": note or "",
        })
    rows.sort(key=lambda r: (r["listing"], r["code"]))

    wb = Workbook()
    ws = wb.active
    ws.title = f"近半年CB(自{since})"
    headers = ["股號", "股名", "發行總量(張)", f"上市櫃後第{NTH_DAY}天成交量(張)",
               "上市櫃日", f"第{NTH_DAY}天日期", "備註"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F3864")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def fmt(d: str | None) -> str:
        return f"{d[:4]}/{d[4:6]}/{d[6:]}" if d else ""

    for r in rows:
        ws.append([r["code"], r["name"], r["total"], r["vol"],
                   fmt(r["listing"]), fmt(r["day6"]), r["note"]])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "#,##0"
    widths = [10, 18, 14, 22, 12, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(out_path)

    print(f"自 {fmt(since)} 起發行 CB 共 {len(rows)} 檔 → {out_path}")
    if dead_dates:
        print(f"已排除 {len(dead_dates)} 個全市場零成交的假日殘欄: "
              + ", ".join(fmt(d) for d in dead_dates))
    miss = [r for r in rows if r["note"]]
    if miss:
        print("需留意:")
        for r in miss:
            print(f"  {r['code']} {r['name']}: {r['note']}")


if __name__ == "__main__":
    main()
