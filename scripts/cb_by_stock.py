# -*- coding: utf-8 -*-
"""目前資料範圍內的 CB,依「個股」彙整成一列:
   股號 | 股名 | 股本(億) | CB1發行總額(億) | CB2… | 備註(2027年到期或強制贖回)

用法: python scripts/cb_by_stock.py [-o 輸出.xlsx]

資料來源:
  data/all-data.json
    cbDailyTrading                  → 資料範圍內有哪些 CB (母體)
    cbasCalendar.issuedInfo         → 發行總額(actualTotal,百萬元)、到期日  [主]
    yuantaReport.basicInfo          → 同上,補統一 CBAS 已移除的已贖回/到期 CB  [備]
    yuantaReport.callRights         → 已行使贖回權 (強制贖回) 的 CB
  data/stock_capital.json           → 股本(億元),MOPS t187ap03 實收資本額

備註欄規則 (使用者指定):
  到期日落在 2027 年 → 「{CB名}{YYYYMMDD}到期」
  出現在 callRights  → 「{CB名}強制贖回」
  同一檔股票多筆以「、」串接。
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from cb_day6_volume import drop_holiday_columns, load_volume_series
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

ROOT = Path(__file__).resolve().parent.parent
ALL_DATA = ROOT / "data" / "all-data.json"
CAPITAL = ROOT / "data" / "stock_capital.json"
MATURITY_YEAR = "2027"
STALE_LOOKBACK = 20   # 近 N 個交易日都沒成交,又查不到發行總額 → 視為已下市殘留列


def _amount_e8(rec: dict | None) -> float | None:
    """actualTotal 單位是百萬元 → 億元。"""
    if not rec:
        return None
    v = rec.get("actualTotal")
    return round(v / 100, 4) if v else None


def main() -> None:
    args = sys.argv[1:]
    out_path = ROOT / "CB依個股彙整.xlsx"
    if "-o" in args:
        i = args.index("-o")
        out_path = Path(args[i + 1])

    all_data = json.loads(ALL_DATA.read_text(encoding="utf-8"))
    issued = all_data["cbasCalendar"]["issuedInfo"]
    basic = all_data["yuantaReport"]["basicInfo"]
    cap_data = json.loads(CAPITAL.read_text(encoding="utf-8"))["data"]

    # 強制贖回:從公告主旨抓「代碼：XXXXX」
    called: set[str] = set()
    for r in all_data["yuantaReport"]["callRights"]:
        m = re.search(r"代碼：(\w+)", r.get("subject", ""))
        if m:
            called.add(m.group(1))

    # 母體 = cbDailyTrading 裡的 CB (去重,保留出現順序)
    cb_codes: list[str] = []
    cb_names: dict[str, str] = {}
    for row in all_data["cbDailyTrading"][1:]:
        code = row[0]
        if not code:
            continue
        if code not in cb_names:
            cb_codes.append(code)
        cb_names.setdefault(code, row[1])

    # 判斷殘留列:來源同時混了 5 碼舊代號與 6 碼現行代號 (交換公司債),
    # 舊列早就沒成交、也查不到發行資料。留在主表只會讓欄位出現空洞。
    dates, series = load_volume_series(all_data)
    dates, _ = drop_holiday_columns(dates, series)
    recent = dates[-STALE_LOOKBACK:]

    def traded_recently(code: str) -> bool:
        s = series.get(code, {})
        return any(str(s.get(d, "0")).strip() not in ("", "-", "0") for d in recent)

    by_stock: dict[str, list[dict]] = defaultdict(list)
    stale: list[dict] = []
    for code in cb_codes:
        rec = issued.get(code)
        src = "統一CBAS"
        if _amount_e8(rec) is None:
            rec, src = basic.get(code), "元大basicInfo"
        amount = _amount_e8(rec)
        if amount is None:
            src = "查無"
        name = (rec or {}).get("cbName") or cb_names.get(code) or ""
        maturity = (rec or {}).get("maturityDate") or ""
        item = {
            "cbCode": code,
            "cbName": name,
            "amount": amount,
            "maturity": maturity,
            "called": code in called,
            "source": src,
            "stockCode": code[:4],
        }
        if amount is None and not traded_recently(code):
            item["source"] = "已下市(不計入主表)"
            stale.append(item)
            continue
        by_stock[code[:4]].append(item)

    max_cb = max(len(v) for v in by_stock.values())

    rows = []
    for stock_code in sorted(by_stock):
        cbs = sorted(by_stock[stock_code], key=lambda c: c["cbCode"])
        info = cap_data.get(stock_code) or {}
        notes = []
        for c in cbs:
            if c["called"]:
                notes.append(f"{c['cbName']}強制贖回")
            elif c["maturity"].startswith(MATURITY_YEAR):
                notes.append(f"{c['cbName']}{c['maturity'].replace('-', '')}到期")
            elif c["amount"] is None:
                notes.append(f"{c['cbName']}發行總額查無")
        rows.append({
            "code": stock_code,
            "name": info.get("name") or cbs[0]["cbName"],
            "capital": info.get("capital"),
            "amounts": [c["amount"] for c in cbs],
            "note": "、".join(notes),
            "cbs": cbs,
        })

    wb = Workbook()

    # ── 主表 ──
    ws = wb.active
    ws.title = "依個股彙整"
    headers = (["股號", "股名", "股本(億)"]
               + [f"CB{i}發行總額(億)" for i in range(1, max_cb + 1)]
               + [f"備註({MATURITY_YEAR}年到期或強制贖回)"])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        amounts = r["amounts"] + [None] * (max_cb - len(r["amounts"]))
        ws.append([r["code"], r["name"], r["capital"], *amounts, r["note"]])

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3 + max_cb):
        for cell in row:
            cell.number_format = "#,##0.00"
    for i, w in enumerate([10, 14, 12] + [15] * max_cb + [46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ── 明細表 (逐檔 CB,方便核對主表數字從哪來) ──
    ws2 = wb.create_sheet("CB明細")
    h2 = ["股號", "股名", "CB代號", "CB名", "發行總額(億)", "到期日", "強制贖回", "來源"]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    name_of = {r["code"]: r["name"] for r in rows}
    detail = [(r["code"], c) for r in rows for c in r["cbs"]]
    detail += [(c["stockCode"], c) for c in stale]
    for stock_code, c in sorted(detail, key=lambda x: (x[0], x[1]["cbCode"])):
        ws2.append([stock_code, name_of.get(stock_code, ""), c["cbCode"], c["cbName"],
                    c["amount"], c["maturity"],
                    "Y" if c["called"] else "", c["source"]])
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"
    for i, w in enumerate([10, 14, 10, 16, 14, 12, 10, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{ws2.max_row}"

    wb.save(out_path)

    total_cb = sum(len(v) for v in by_stock.values())
    no_amt = [c for r in rows for c in r["cbs"] if c["amount"] is None]
    no_cap = [r for r in rows if r["capital"] is None]
    print(f"{len(rows)} 檔股票 / {total_cb} 檔 CB,最多 {max_cb} 檔/股 → {out_path}")
    print(f"  已下市殘留列排除 {len(stale)} 檔 (仍列在 CB明細 分頁)")
    print(f"  發行總額查無 {len(no_amt)} 檔 CB;股本查無 {len(no_cap)} 檔股票"
          + (f" ({', '.join(r['code'] for r in no_cap)})" if no_cap else ""))
    from collections import Counter
    print("  發行總額來源:", dict(Counter(c["source"] for r in rows for c in r["cbs"])))


if __name__ == "__main__":
    main()
