"""統一 CBAS 日曆事件:已發行CB資料.xlsx + 預計發行CB資料.xlsx → calendar events。

來源 Drive folder (統一CBAS日曆),使用者每日把統一 CBAS 平台匯出的兩份 xlsx
上傳進來,檔名形如:
  已發行CB資料_YYYYMMDD.xlsx
  預計發行CB資料_YYYYMMDD.xlsx

本模組挑檔名日期最大那組,解析成 calendar event 清單。

事件型別 (type):
  issue        CB發行日       (已發行.發行日期 / 預計發行.近期掛牌.掛牌日)
  aso          CB拆解日       (預計發行.近期掛牌.拆解日)
  bookbuilding CB詢圈期間     (預計發行.近期掛牌.詢圈/競拍,range)
  auction      CB競拍期間     (預計發行.近期掛牌.詢圈/競拍,range)
  board        董事會公告日   (預計發行.董事會公告/近期生效.公告日)
  maturity     CB到期日       (已發行.到期日)
  putback      CB賣回日       (已發行.最新賣回日)
  forcedRedeem 強制贖回日     (已發行.強制贖回日)
  resetConv    重設轉換日     (已發行.重設轉換日)

point event : {date, type, cbCode, cbName}
range event : {date, endDate, type, cbCode, cbName}

跑法 (smoke,需 .env 內 GOOGLE_CREDENTIALS):
  python -m lib.cbas_calendar
"""

from __future__ import annotations

import datetime as _dt
import io
import re
import sys
from typing import Any, Optional

import requests
import urllib3

# Windows console UTF-8
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

# 統一CBAS日曆 folder (API 抓不到時的 fallback 來源)
DEFAULT_FOLDER_ID = "1qAauB30BsCZ2_dHMJ_3qlTfr5IhK0Q2s"

# 統一 CBAS 平台下載 API (公開 GET,無需登入)
API_ISSUED = "https://cbas16889.pscnet.com.tw/api/MiDownloadExcel/GetExcel_IssuedCB"
API_RECENTLY = "https://cbas16889.pscnet.com.tw/api/MiDownloadExcel/GetExcel_RecentlyCB"
_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)
# 統一憑證缺 Subject Key Identifier,Python 嚴格驗證會擋 → verify=False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

_DATE_IN_NAME = re.compile(r"(\d{8})")
# 詢圈/競拍 文字:M/D-M/D詢圈 或 M/D-M/D競拍
_PERIOD_RE = re.compile(
    r"(\d{1,2})/(\d{1,2})\s*[-~]\s*(\d{1,2})/(\d{1,2})\s*(詢圈|競拍)"
)


# ── 日期正規化 ────────────────────────────────────────────────────────
def _to_iso(v: Any) -> Optional[str]:
    """xlsx cell → 'YYYY-MM-DD';無法解析 / 年份不合理回 None。"""
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        y = v.year
        return v.strftime("%Y-%m-%d") if 2000 <= y <= 2100 else None
    s = str(v).strip()
    if not s or s in ("-", "—", "未定", "N/A"):
        return None
    # 2026/05/11 或 2026/5/18 或 2026-05-11
    m = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # 排除 Excel epoch 殘值 (1899/1900) 與不合理年份
        if not (2000 <= y <= 2100):
            return None
        try:
            return _dt.date(y, mo, d).isoformat()
        except ValueError:
            return None
    return None


def _parse_period(text: Any, list_iso: Optional[str]) -> Optional[dict]:
    """'4/23-4/27競拍' → {start, end, type};無法解析回 None。

    年份:詢圈/競拍 一定在掛牌日之前,所以用掛牌日的年份;若推算結果晚於
    掛牌日 (跨年) 則減一年。
    """
    if not text:
        return None
    m = _PERIOD_RE.search(str(text))
    if not m:
        return None
    m1, d1, m2, d2, kind = m.groups()
    base_year = int(list_iso[:4]) if list_iso else _dt.date.today().year
    try:
        start = _dt.date(base_year, int(m1), int(d1))
        end = _dt.date(base_year, int(m2), int(d2))
    except ValueError:
        return None
    # 跨年修正:期間應早於掛牌日
    if list_iso:
        list_d = _dt.date.fromisoformat(list_iso)
        if start > list_d:
            start = start.replace(year=base_year - 1)
        if end > list_d:
            end = end.replace(year=base_year - 1)
    # 來源偶有 end<start 的筆誤 (如 5/9-5/8),取 min/max
    if end < start:
        start, end = end, start
    etype = "auction" if kind == "競拍" else "bookbuilding"
    return {"start": start.isoformat(), "end": end.isoformat(), "type": etype}


# ── 欄位定位 (用 header 名稱,避免欄序變動) ───────────────────────────
def _header_map(header_row: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, cell in enumerate(header_row or ()):
        key = str(cell or "").strip()
        if key:
            out[key] = i
    return out


def _cell(row: tuple, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _valid_cb(code: Any) -> bool:
    s = str(code or "").strip()
    return s.isdigit() and 4 <= len(s) <= 6


# ── 解析:已發行CB資料 ───────────────────────────────────────────────
def parse_issued(blob: bytes) -> list[dict]:
    from openpyxl import load_workbook

    # 不用 read_only:統一匯出的 xlsx 缺 dimension metadata,
    # read_only 模式會誤判成只有 1 列
    wb = load_workbook(io.BytesIO(blob), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    h = _header_map(rows[0])
    c_code = h.get("債券代號", 0)
    c_name = h.get("標的債券", 1)
    field_cols = {
        "issue": h.get("發行日期"),
        "maturity": h.get("到期日"),
        "putback": h.get("最新賣回日"),
        "forcedRedeem": h.get("強制贖回日"),
        "resetConv": h.get("重設轉換日"),
    }
    events: list[dict] = []
    for row in rows[1:]:
        code = str(_cell(row, c_code) or "").strip()
        if not _valid_cb(code):
            continue
        name = str(_cell(row, c_name) or "").strip()
        for etype, col in field_cols.items():
            iso = _to_iso(_cell(row, col))
            if iso:
                events.append({"date": iso, "type": etype,
                               "cbCode": code, "cbName": name})
    return events


# ── 解析:預計發行CB資料 ─────────────────────────────────────────────
def parse_planned(blob: bytes) -> list[dict]:
    from openpyxl import load_workbook

    # 不用 read_only:統一匯出的 xlsx 缺 dimension metadata,
    # read_only 模式會誤判成只有 1 列
    wb = load_workbook(io.BytesIO(blob), data_only=True)
    events: list[dict] = []

    # 近期掛牌:掛牌日(issue) + 拆解日(aso) + 詢圈/競拍(range)
    if "近期掛牌" in wb.sheetnames:
        rows = list(wb["近期掛牌"].iter_rows(values_only=True))
        if rows:
            h = _header_map(rows[0])
            c_code = h.get("CB代號", 1)
            c_name = h.get("CB名稱", 2)
            c_list = h.get("掛牌日")
            c_aso = h.get("拆解日")
            c_period = h.get("詢圈/競拍")
            for row in rows[1:]:
                code = str(_cell(row, c_code) or "").strip()
                if not _valid_cb(code):
                    continue
                name = str(_cell(row, c_name) or "").strip()
                list_iso = _to_iso(_cell(row, c_list))
                if list_iso:
                    events.append({"date": list_iso, "type": "issue",
                                   "cbCode": code, "cbName": name})
                aso_iso = _to_iso(_cell(row, c_aso))
                if aso_iso:
                    events.append({"date": aso_iso, "type": "aso",
                                   "cbCode": code, "cbName": name})
                period = _parse_period(_cell(row, c_period), list_iso)
                if period:
                    events.append({"date": period["start"],
                                   "endDate": period["end"],
                                   "type": period["type"],
                                   "cbCode": code, "cbName": name})

    # 近期生效 + 董事會公告:公告日(board)
    for sheet_name in ("近期生效", "董事會公告"):
        if sheet_name not in wb.sheetnames:
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        h = _header_map(rows[0])
        c_code = h.get("CB代號", 1)
        c_name = h.get("CB名稱", 2)
        c_board = h.get("公告日")
        for row in rows[1:]:
            code = str(_cell(row, c_code) or "").strip()
            if not _valid_cb(code):
                continue
            name = str(_cell(row, c_name) or "").strip()
            board_iso = _to_iso(_cell(row, c_board))
            if board_iso:
                events.append({"date": board_iso, "type": "board",
                               "cbCode": code, "cbName": name})

    return events


def _dedup(events: list[dict]) -> list[dict]:
    """以 (cbCode,type,date) 去重,保留首見。"""
    seen: set = set()
    out: list[dict] = []
    for e in events:
        key = (e["cbCode"], e["type"], e["date"])
        if key in seen:
            continue
        seen.add(key)
        out.append(e)
    out.sort(key=lambda e: (e["date"], e["type"], e["cbCode"]))
    return out


# ── 直接從統一 CBAS API 下載 (主要來源) ─────────────────────────────
def _api_get(url: str) -> tuple[Optional[str], bytes]:
    """GET 一個 CBAS 下載 API,回傳 (date_str|None, xlsx_bytes)。"""
    r = requests.get(
        url, timeout=60, verify=False,
        headers={"User-Agent": _UA,
                 "Referer": "https://cbas16889.pscnet.com.tw/",
                 "Accept": "*/*"},
    )
    r.raise_for_status()
    if r.content[:2] != b"PK":
        raise RuntimeError(f"回應不是 xlsx (前 8 bytes: {r.content[:8]!r})")
    # 檔名形如 20260521_已發行CB資料.xlsx → 取開頭 8 碼日期
    cd = r.headers.get("content-disposition", "")
    m = _DATE_IN_NAME.search(cd)
    return (m.group(1) if m else None), r.content


def fetch_from_api() -> tuple[str, bytes, bytes]:
    """從統一 CBAS API 直接下載兩份 xlsx。回傳 (date_str, issued, planned)。"""
    d1, issued = _api_get(API_ISSUED)
    d2, planned = _api_get(API_RECENTLY)
    date_str = d1 or d2 or _dt.datetime.now(
        _dt.timezone(_dt.timedelta(hours=8))).strftime("%Y%m%d")
    return date_str, issued, planned


# ── Drive 取最新一對 xlsx (API 失敗時 fallback) ──────────────────────
def fetch_latest_pair(folder_id: str = DEFAULT_FOLDER_ID
                      ) -> tuple[str, bytes, bytes]:
    """回傳 (date_str, issued_blob, planned_blob)。失敗拋 RuntimeError。"""
    from lib import drive  # 重用既有 service account 連線

    svc = drive._get_service()
    res = svc.files().list(
        q=f"'{folder_id}' in parents and trashed = false",
        fields="files(id,name)",
        pageSize=200,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    files = res.get("files", [])
    if not files:
        raise RuntimeError(f"統一CBAS日曆 folder {folder_id} 沒有任何檔案")

    def _date(name: str) -> str:
        m = _DATE_IN_NAME.search(name or "")
        return m.group(1) if m else ""

    issued = [f for f in files if "已發行" in f["name"]]
    planned = [f for f in files if "預計發行" in f["name"]]
    if not issued or not planned:
        raise RuntimeError(
            f"folder 內缺『已發行』或『預計發行』xlsx (issued={len(issued)} "
            f"planned={len(planned)})"
        )
    issued.sort(key=lambda f: _date(f["name"]), reverse=True)
    planned.sort(key=lambda f: _date(f["name"]), reverse=True)
    latest_issued = issued[0]
    latest_planned = planned[0]
    date_str = _date(latest_issued["name"]) or _date(latest_planned["name"])

    def _download(file_id: str) -> bytes:
        from googleapiclient.http import MediaIoBaseDownload
        req = svc.files().get_media(fileId=file_id, supportsAllDrives=True)
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        return buf.getvalue()

    return (date_str,
            _download(latest_issued["id"]),
            _download(latest_planned["id"]))


def fetch_and_parse(folder_id: str = DEFAULT_FOLDER_ID) -> dict[str, Any]:
    """主入口:抓最新 xlsx → 解析 → {reportDate, events, source}。

    優先直接打統一 CBAS API (最即時);失敗才退回 Drive 上 fetch_stocks
    備份的 xlsx。
    """
    source = "api"
    try:
        date_str, issued_blob, planned_blob = fetch_from_api()
    except Exception as e:  # noqa: BLE001
        print(f"⚠️ CBAS API 下載失敗,改讀 Drive 備份: {e}",
              file=sys.stderr, flush=True)
        date_str, issued_blob, planned_blob = fetch_latest_pair(folder_id)
        source = "drive"
    events = parse_issued(issued_blob) + parse_planned(planned_blob)
    return {"reportDate": date_str, "events": _dedup(events), "source": source}


# ── Smoke test (走 API,不需任何憑證) ────────────────────────────────
def _smoke() -> int:
    result = fetch_and_parse()
    events = result["events"]
    print(f"source={result['source']}  reportDate={result['reportDate']}  "
          f"events={len(events)}")
    from collections import Counter
    for t, n in Counter(e["type"] for e in events).most_common():
        print(f"  {t:14s} {n}")
    print("\n前 12 筆:")
    for e in events[:12]:
        rng = f" ~ {e['endDate']}" if e.get("endDate") else ""
        print(f"  {e['date']}{rng}  {e['type']:13s} {e['cbCode']} {e['cbName']}")
    return 0


if __name__ == "__main__":
    sys.exit(_smoke())
