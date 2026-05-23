"""個股狀態 sheet 抓取 (VCP / 三線開花) → dict[stock_code, details]。

兩份 Google Sheet 結構共同點:
  - 公開 (任何人皆可檢視)
  - 每天新增一個工作表,sheet name 為 YYMMDD (如 '260505')
  - A 欄 = 股票代號, B/C = 股名/公司名, D-F = 產業分類, 後續為各指標欄位

抓法:
  下載整份 xlsx (export?format=xlsx),依 sheet name 排序找最新
  -- 比 gviz 的 gid 路線穩定,不需要每天更新 gid

跑法 (smoke):
  python -m lib.status_sheets

輸出範例:
  {
    "vcp":     {"date": "260505", "stocks": {"2423": {...}, ...}},
    "sanxian": {"date": "260505", "stocks": {"2342": {...}, ...}}
  }
"""

from __future__ import annotations

import io
import re
import sys
from typing import Any, Optional

import requests

# Windows console UTF-8
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

XLSX_URL_TPL = (
    "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
)
TIMEOUT = 60

# 兩份固定的狀態 sheet
SOURCES: dict[str, dict[str, Any]] = {
    "vcp": {
        "name": "VCP",
        "sheet_id": "1zJ0st8vbhrINdBvfi0-U3gFcbqeFHHHsUEEULM3kDSQ",
        # 標頭: 代碼 / 股名 / 公司名 / 產業1 / 產業2 / 產品組合 /
        #        近20日股價漲幅 / 大盤淨空 / 連續淨空
        "fields": {
            6: "gain20",       # 近20日股價漲幅
            7: "marketShort",  # 大盤淨空 (O / 空)
            8: "consecShort",  # 連續淨空 (O / 空)
        },
    },
    "sanxian": {
        "name": "三線開花",
        "sheet_id": "1HinNbehTVBoBdZV6FqQRhld-WOWkSJvfjd0e1ndDvxo",
        # 標頭: 代碼 / 股名 / 公司名 / 產業1 / 產業2 / 產品組合 /
        #        收盤股價 / 55日內最高價 / 差距比
        "fields": {
            6: "close",       # 收盤股價
            7: "high55",      # 55日內最高價
            8: "diffPct",     # 差距比
        },
    },
}

# YYMMDD pattern for sheet name (e.g. 260505)
DATE_RE = re.compile(r"^(\d{6})$")


def _pick_latest_sheet(sheet_names: list[str]) -> Optional[str]:
    """從 sheet 名稱中挑最新一個。

    優先挑 YYMMDD 格式排序最大的;若全不符,退回原始第 0 個 (Google Sheet 預設第一個)。
    """
    dated = [n for n in sheet_names if DATE_RE.match(n)]
    if dated:
        return max(dated)
    return sheet_names[0] if sheet_names else None


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # 整數浮點 → 去 .0;其餘保留 2 位
        if v.is_integer():
            return str(int(v))
        return f"{v:.4f}".rstrip("0").rstrip(".")
    return str(v).strip()


def _is_valid_code(s: str) -> bool:
    return s.isdigit() and 4 <= len(s) <= 6


def fetch_one(key: str) -> dict[str, Any]:
    """抓單一 sheet → {date, stocks, sheetsScanned}。失敗 raise RuntimeError。

    掃 xlsx 內所有 YYMMDD 工作表計算每檔的:
      - streak: 從最新往回連續上榜天數 (今天首上榜=1)
      - total : 在歷史範圍內累計上榜天數
    """
    spec = SOURCES[key]
    url = XLSX_URL_TPL.format(sheet_id=spec["sheet_id"])
    r = requests.get(url, timeout=TIMEOUT, allow_redirects=True)
    r.raise_for_status()
    if not r.content:
        raise RuntimeError(f"{spec['name']}: xlsx 內容為空")

    # 延後 import,讓 import error 訊息直接顯示在 fetch 階段
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f"{spec['name']}: 解析 xlsx 失敗: {e}") from e

    # 全部 YYMMDD 工作表,降冪 (最新優先)
    dated = sorted(
        (s for s in wb.sheetnames if DATE_RE.match(s)),
        reverse=True,
    )
    if not dated:
        raise RuntimeError(f"{spec['name']}: 無任何 YYMMDD 工作表")
    latest = dated[0]

    # 第一遍:掃所有工作表,建出每檔股票的「出現日期集合」
    appearance: dict[str, set[str]] = {}
    for tab in dated:
        for row in wb[tab].iter_rows(values_only=True):
            if not row:
                continue
            code = _stringify(row[0])
            if not _is_valid_code(code):
                continue
            appearance.setdefault(code, set()).add(tab)

    # 第二遍:最新工作表 → 當前明細,加 streak / total
    fields = spec["fields"]
    stocks: dict[str, dict[str, Any]] = {}
    for row in wb[latest].iter_rows(values_only=True):
        if not row:
            continue
        code = _stringify(row[0])
        if not _is_valid_code(code):
            continue
        item: dict[str, Any] = {}
        for col_idx, field_name in fields.items():
            if col_idx < len(row):
                val = _stringify(row[col_idx])
                if val:
                    item[field_name] = val
        appear_set = appearance.get(code, set())
        streak = 0
        for tab in dated:
            if tab in appear_set:
                streak += 1
            else:
                break
        item["streak"] = streak
        item["total"] = len(appear_set)
        stocks[code] = item

    return {"date": latest, "stocks": stocks, "sheetsScanned": len(dated)}


def fetch_all() -> dict[str, dict[str, Any]]:
    """抓兩份 sheet,個別失敗不會影響另一份。回傳 dict 會省略失敗的 key。

    呼叫端可從回傳 dict 的 key 數量得知成功與否。
    """
    out: dict[str, dict[str, Any]] = {}
    for key in SOURCES:
        try:
            out[key] = fetch_one(key)
        except Exception as e:  # noqa: BLE001
            print(f"⚠️ status_sheets[{key}]: {e}", file=sys.stderr, flush=True)
    return out


# ── Smoke test ────────────────────────────────────────────────────────
def _smoke() -> int:
    for key in SOURCES:
        try:
            res = fetch_one(key)
            print(f"[{key}] tab={res['date']} scanned={res['sheetsScanned']} "
                  f"stocks={len(res['stocks'])}")
            # streak 最長前 5 名
            top = sorted(res["stocks"].items(),
                         key=lambda kv: kv[1].get("streak", 0), reverse=True)[:5]
            for code, details in top:
                s, t = details.get("streak"), details.get("total")
                print(f"  {code}: streak={s} total={t} {details}")
        except Exception as e:
            print(f"[{key}] FAIL: {e}")
            return 1
    return 0


if __name__ == "__main__":
    sys.exit(_smoke())
