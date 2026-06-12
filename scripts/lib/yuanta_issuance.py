"""元大證券「CB發行案件彙整」xlsx → CB 日曆補充來源 + 與統一 CBAS 核對。

來源:Google Drive folder (CB發行案件-元大),挑檔名
  CB發行案件彙整-元大債券{ROC7}.xlsx   (ROC7 = 民國年月日,例 1150612 = 2026/06/12)
日期最大那份。元大每天更新,常比統一 CBAS 平台「更早」收錄新發行案,
並補上掛牌日 / 拆解日,適合拿來補強 + 核對統一 CBAS 日曆。

只解析「發行案件」分頁 (近期/進行中的發行案)。欄位 (header 在第 2 列):
  代號(stockCode, 無表頭) | 詢圈/競拍 | 標的代號(CB) | 發行標的(CB名) |
  TCRI/擔保 | 發行量 | 主辦券商 | 送件日 | 生效日 | 詢圈/投標期間 |
  溢價率 | 轉換價 | 掛牌日 | 可拆解選擇權日 | 承銷價格 | 年期 |
  賣回條件 | 股本 | 股價 | 60天波動率 | 備註

對齊 cbas_calendar 的事件型別,元大可貢獻:
  issue        掛牌日
  aso          可拆解選擇權日 (欄位常夾雜文字,如「額度有限2026/6/4」→ 取日期)
  bookbuilding 詢圈/投標期間 (kind=詢圈)
  auction      詢圈/投標期間 (kind=競拍)

注意:部分早期案件 (僅董事會階段) 欄位左移、掛牌/拆解/期間皆空,
這類 row 不會產生任何日期事件,只會在核對結果列為「僅元大、尚無日曆日期」。

跑法 (smoke,需 .env 內 GOOGLE_CREDENTIALS):
  python -m lib.yuanta_issuance
"""

from __future__ import annotations

import datetime as _dt
import io
import re
import sys
from typing import Any, Optional

# Windows console UTF-8
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

# CB發行案件-元大 Drive folder
DEFAULT_FOLDER_ID = "1wZcIAX6Jaw3UcZs6SRwIXAsAwXEHqmYx"
FILE_NAME_RE = re.compile(r"CB發行案件彙整-元大債券(\d{7})\.xlsx$")
SHEET_NAME = "發行案件"

# 任意位置的西元日期 (允許前後夾雜文字,如「額度有限2026/6/4」)
_DATE_ANYWHERE = re.compile(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})")
# 詢圈/投標期間:M/D-M/D 或 M/D~M/D 或 M/D/M/D (來源偶用 / 當分隔)
_PERIOD_RE = re.compile(r"(\d{1,2})/(\d{1,2})\s*[-~/]\s*(\d{1,2})/(\d{1,2})")


# ── 日期工具 ──────────────────────────────────────────────────────────
def _roc7_to_iso(roc7: str) -> str:
    """民國 7 碼 '1150612' → '2026-06-12';失敗回原字串。"""
    try:
        y = int(roc7[:3]) + 1911
        return _dt.date(y, int(roc7[3:5]), int(roc7[5:7])).isoformat()
    except (ValueError, IndexError):
        return roc7


def _search_date(v: Any) -> Optional[str]:
    """cell → 'YYYY-MM-DD';datetime 直接取,字串則搜尋內含日期。無/不合理回 None。"""
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d") if 2000 <= v.year <= 2100 else None
    m = _DATE_ANYWHERE.search(str(v))
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not (2000 <= y <= 2100):
        return None
    try:
        return _dt.date(y, mo, d).isoformat()
    except ValueError:
        return None


def _parse_period(text: Any, base_iso: Optional[str], kind: str) -> Optional[dict]:
    """'5/13-5/15' + kind → {start, end, type};無法解析回 None。

    年份:期間在掛牌日之前,優先用 base_iso (掛牌日/生效日/報告日) 的年份;
    若推算結果晚於 base_iso 視為跨年,減一年。
    """
    if not text:
        return None
    m = _PERIOD_RE.search(str(text))
    if not m:
        return None
    m1, d1, m2, d2 = (int(x) for x in m.groups())
    base_year = int(base_iso[:4]) if base_iso else _dt.date.today().year
    try:
        start = _dt.date(base_year, m1, d1)
        end = _dt.date(base_year, m2, d2)
    except ValueError:
        return None
    if base_iso:
        base_d = _dt.date.fromisoformat(base_iso)
        if start > base_d:
            start = start.replace(year=base_year - 1)
        if end > base_d:
            end = end.replace(year=base_year - 1)
    if end < start:
        start, end = end, start
    etype = "auction" if "競拍" in kind else "bookbuilding"
    return {"start": start.isoformat(), "end": end.isoformat(), "type": etype}


# ── 欄位定位 ──────────────────────────────────────────────────────────
def _header_map(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    """找含「標的代號」的表頭列,回 (列索引, {表頭: 欄索引})。找不到回 (-1, {})。"""
    for r in range(min(6, len(rows))):
        row = rows[r] or ()
        for i, cell in enumerate(row):
            if str(cell or "").strip() == "標的代號":
                out = {str(c or "").strip(): j
                       for j, c in enumerate(row) if str(c or "").strip()}
                return r, out
    return -1, {}


def _cell(row: tuple, idx: Optional[int]) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _str(v: Any) -> str:
    s = str(v if v is not None else "").strip()
    return "" if s in ("-", "—", "未定", "N/A") else s


def _valid_cb(code: Any) -> bool:
    s = str(code or "").strip()
    return s.isdigit() and 4 <= len(s) <= 6


# ── 解析「發行案件」分頁 ───────────────────────────────────────────────
def parse_issuance(blob: bytes, report_date: str = "") -> dict[str, Any]:
    """回傳 {reportDate, entries, events}。

    entries — 每檔的原始欄位 (給核對 / debug)
    events  — issue / aso / bookbuilding / auction 事件,結構同 cbas_calendar
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(blob), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"xlsx 缺「{SHEET_NAME}」分頁 (有: {wb.sheetnames})")
    rows = list(wb[SHEET_NAME].iter_rows(values_only=True))
    hr, h = _header_map(rows)
    if hr < 0:
        raise RuntimeError("「發行案件」找不到表頭 (無「標的代號」欄)")

    c_cb = h.get("標的代號", 2)
    c_name = h.get("發行標的", 3)
    c_kind = h.get("詢圈/競拍", 1)
    c_list = h.get("掛牌日")
    c_aso = h.get("可拆解選擇權日")
    c_period = h.get("詢圈/投標期間")
    c_eff = h.get("生效日")
    c_file = h.get("送件日")

    entries: list[dict] = []
    events: list[dict] = []
    for row in rows[hr + 1:]:
        cb = str(_cell(row, c_cb) or "").strip()
        if not _valid_cb(cb):
            continue
        name = _str(_cell(row, c_name))
        kind = _str(_cell(row, c_kind))
        stock = str(_cell(row, 0) or "").strip()  # 代號欄無表頭,固定第 0 欄
        list_iso = _search_date(_cell(row, c_list))
        aso_iso = _search_date(_cell(row, c_aso))
        eff_iso = _search_date(_cell(row, c_eff))
        file_iso = _search_date(_cell(row, c_file))
        period_raw = _str(_cell(row, c_period))

        entry = {
            "cbCode": cb, "cbName": name,
            "stockCode": stock[:4] if stock[:4].isdigit() else cb[:4],
            "kind": kind, "listDate": list_iso, "asoDate": aso_iso,
            "effDate": eff_iso, "filingDate": file_iso, "period": period_raw,
            "underwriter": _str(_cell(row, h.get("主辦券商"))),
            "tcriGuarantee": _str(_cell(row, h.get("TCRI/擔保"))),
        }
        entries.append(entry)

        if list_iso:
            events.append({"date": list_iso, "type": "issue",
                           "cbCode": cb, "cbName": name})
        if aso_iso:
            events.append({"date": aso_iso, "type": "aso",
                           "cbCode": cb, "cbName": name})
        # 期間事件:年份基準優先掛牌日,否則用報告日 (發行案件皆為近期案,
        # 報告日比可能過期的生效日可靠;_parse_period 內含跨年 (12月→隔年1月) 修正)
        base_iso = list_iso or (_roc7_to_iso(report_date) if report_date else None)
        period = _parse_period(period_raw, base_iso, kind)
        if period:
            events.append({"date": period["start"], "endDate": period["end"],
                           "type": period["type"], "cbCode": cb, "cbName": name})

    return {"reportDate": _roc7_to_iso(report_date) if report_date else "",
            "reportDateRoc": report_date, "entries": entries, "events": events}


# ── Drive 取最新檔 ───────────────────────────────────────────────────
def fetch_latest_xlsx(folder_id: str = DEFAULT_FOLDER_ID) -> tuple[str, bytes]:
    """撈 'CB發行案件彙整-元大債券{ROC7}.xlsx' 最大日期那份 → (roc7, bytes)。"""
    from lib import drive

    svc = drive._get_service()
    files = svc.files().list(
        q=f"'{folder_id}' in parents and trashed = false",
        fields="files(id,name)", pageSize=200,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute().get("files", [])

    cands: list[tuple[str, dict]] = []
    for f in files:
        m = FILE_NAME_RE.search(f.get("name", ""))
        if m:
            cands.append((m.group(1), f))
    if not cands:
        raise RuntimeError(
            f"folder {folder_id} 內找不到「CB發行案件彙整-元大債券YYYMMDD.xlsx」")
    cands.sort(key=lambda x: x[0], reverse=True)
    roc7, target = cands[0]

    from googleapiclient.http import MediaIoBaseDownload
    req = svc.files().get_media(fileId=target["id"], supportsAllDrives=True)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return roc7, buf.getvalue()


def fetch_and_parse(folder_id: str = DEFAULT_FOLDER_ID) -> dict[str, Any]:
    """一站式:Drive 取最新檔 → 解析「發行案件」。"""
    roc7, blob = fetch_latest_xlsx(folder_id)
    return parse_issuance(blob, report_date=roc7)


# ── 與統一 CBAS 日曆核對 + 補充 ───────────────────────────────────────
def crosscheck_and_merge(cal: dict, yuanta: dict) -> dict[str, Any]:
    """以統一 CBAS (cal) 為主,用元大 (yuanta) 補充缺漏事件並標記差異。

    *就地修改* cal['events']:把統一缺的 (cbCode,type) 事件補進去
    (每筆標 source='yuanta'),統一已有同型別則保留統一、不覆蓋。

    回傳核對摘要 dict;同時掛在 cal['yuantaCrosscheck']。
    """
    from collections import defaultdict
    from lib.cbas_calendar import _dedup

    base_events = cal.get("events") or []
    # 統一既有事件索引:cbCode → {type: set(date)}
    idx: dict[str, dict[str, set]] = defaultdict(lambda: defaultdict(set))
    for e in base_events:
        idx[e["cbCode"]][e["type"]].add(e["date"])
    present_cb = set(idx.keys())

    # 真正產出事件 (掛牌/拆解/期間可解析) 的 CB,才算「已有日曆日期」
    evented_cb = {e["cbCode"] for e in (yuanta.get("events") or [])}

    supplements: list[dict] = []
    mismatches: list[dict] = []
    consistent = 0
    only_yuanta: list[dict] = []

    for ev in yuanta.get("events") or []:
        cb, etype, date = ev["cbCode"], ev["type"], ev["date"]
        existing = idx[cb].get(etype)
        if existing:
            if date in existing:
                consistent += 1
            else:
                mismatches.append({
                    "cbCode": cb, "cbName": ev.get("cbName", ""),
                    "type": etype, "yuantaDate": date,
                    "cbasDate": sorted(existing)[0],
                })
            continue  # 統一已有此型別 → 不補
        # 統一缺此型別 → 用元大補
        supp = {**ev, "source": "yuanta"}
        supplements.append(supp)
        idx[cb][etype].add(date)  # 避免元大自身重複補

    # 整檔僅元大有 (統一日曆完全沒這檔事件) 的 CB
    for entry in yuanta.get("entries") or []:
        cb = entry["cbCode"]
        if cb not in present_cb:
            only_yuanta.append({
                "cbCode": cb, "cbName": entry.get("cbName", ""),
                "kind": entry.get("kind", ""),
                "hasDate": cb in evented_cb,
            })

    if supplements:
        cal["events"] = _dedup(base_events + supplements)

    summary = {
        "reportDate": yuanta.get("reportDate", ""),
        "reportDateRoc": yuanta.get("reportDateRoc", ""),
        "yuantaRows": len(yuanta.get("entries") or []),
        "consistent": consistent,
        "supplied": len(supplements),
        "mismatches": mismatches,
        "onlyYuanta": only_yuanta,
    }
    cal["yuantaCrosscheck"] = summary
    return summary


# ── Smoke test ────────────────────────────────────────────────────────
def _smoke() -> int:
    from collections import Counter

    res = fetch_and_parse()
    print(f"reportDate = {res['reportDate']} (ROC {res['reportDateRoc']})")
    print(f"entries    = {len(res['entries'])}")
    print(f"events     = {len(res['events'])}")
    for t, n in Counter(e["type"] for e in res["events"]).most_common():
        print(f"  {t:13s} {n}")
    print("\n前 10 筆事件:")
    for e in res["events"][:10]:
        rng = f" ~ {e['endDate']}" if e.get("endDate") else ""
        print(f"  {e['date']}{rng}  {e['type']:13s} {e['cbCode']} {e['cbName']}")
    return 0


if __name__ == "__main__":
    sys.exit(_smoke())
