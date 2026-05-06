"""元大證選擇權 xlsx → yuantaReport dict (Python port of YuantaReportParser.gs)。

來源:Google Drive folder (YUANTA_REPORT),挑檔名 `元大證選擇權YYYYMMDD.xlsx`
日期最大那份。每天會有新檔,常見 callDate / 轉換價格調整等高警示資訊
都來自這裡。

輸出 dict 與 GAS YuantaReportParser 完全對齊 (給 js/dataProcessor.js
mergeYuantaReport 使用):
  {
    reportDate: "YYYYMMDD",
    quotes:         { cbCode: {tcri, premium, optionExpiry, ...} },
    basicInfo:      { cbCode: {cbName, couponRate, convPrice, ...,
                                resetFormula, callDate} },
    callRights:     [ {stockCode, stockName, reportDate, subject, asoExpiry} ],
    conversionStop: [ {cbCode, cbName, startDate, endDate, reason} ],
    priceAdjust:    [ {type, stockCode, stockName, reportDate, subject, newPrice} ],
    outstanding:    { cbCode: {thisWeek, lastWeek, change, ...} },
    dailyMarket:    { cbCode: {ytp, ytm, eps, vol120, ..., business} }
  }

跑法 (smoke):
  python -m lib.yuanta_report
  → 列出 Drive folder + 解析最新一份印彙總
"""

from __future__ import annotations

import io
import os
import re
import sys
from datetime import datetime
from typing import Any, Optional

# Windows console UTF-8
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

# Drive folder ID:CB選擇權-元大 (xlsx_uploads 子目錄)
DEFAULT_FOLDER_ID = "1xHJ5OEnBxyRKSkQYCOZRpZrUs9cRgyHX"
FILE_NAME_RE = re.compile(r"元大證選擇權(\d{8})\.xlsx$")


# ── tools ────────────────────────────────────────────────────────────
def _str(v: Any) -> str:
    if v is None or v == "":
        return ""
    return str(v).strip()


def _num(v: Any) -> Optional[float]:
    if v is None or v == "" or v == "-" or v == "--":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # Excel numbers come through as float/int directly
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _date_str(v: Any) -> str:
    """日期欄轉 'YYYY-MM-DD',空值回 ''。"""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    s = str(v).strip()
    if s in ("-", "--", ""):
        return ""
    return s


def _find_header_row(rows: list[tuple], keyword: str, scan_rows: int = 10,
                     scan_cols: int = 3) -> int:
    for r in range(min(scan_rows, len(rows))):
        row = rows[r] or ()
        for c in range(min(scan_cols, len(row))):
            if keyword in _str(row[c]):
                return r
    return -1


# ── 各分頁 parser ────────────────────────────────────────────────────
def _parse_quotes(rows: list[tuple]) -> dict[str, Any]:
    """報價單: 'header row 含「可轉債名稱」',資料從 header+2 開始,code 在 row[1]。"""
    hdr = _find_header_row(rows, "可轉債名稱")
    if hdr < 0:
        return {}
    out: dict[str, Any] = {}
    for r in range(hdr + 2, len(rows)):
        row = rows[r] or ()
        code = _str(row[1] if len(row) > 1 else "")
        if not code or not re.match(r"^\d{4,6}$", code):
            continue
        out[code] = {
            "cbName":         _str(row[0]),
            "tcri":           _num(row[2] if len(row) > 2 else None),
            "premium":        _num(row[3] if len(row) > 3 else None),
            "optionExpiry":   _date_str(row[4] if len(row) > 4 else None),
            "putPrice":       _num(row[5] if len(row) > 5 else None),
            "putDate":        _date_str(row[6] if len(row) > 6 else None),
            "remainYears":    _num(row[7] if len(row) > 7 else None),
            "discountRate":   _num(row[8] if len(row) > 8 else None),
            "vol120":         _num(row[15] if len(row) > 15 else None),
            "vol240":         _num(row[16] if len(row) > 16 else None),
            "outstandingPct": _num(row[17] if len(row) > 17 else None),
            "issueAmount":    _num(row[18] if len(row) > 18 else None),
            "industry":       _str(row[19] if len(row) > 19 else ""),
        }
    return out


def _parse_basic_info(rows: list[tuple]) -> dict[str, Any]:
    """基本資料檔: header 是「代號」+「名稱」。"""
    hdr = -1
    for r in range(min(10, len(rows))):
        row = rows[r] or ()
        if (len(row) >= 2 and _str(row[0]) == "代號" and _str(row[1]) == "名稱"):
            hdr = r
            break
    if hdr < 0:
        return {}
    out: dict[str, Any] = {}
    for r in range(hdr + 1, len(rows)):
        row = rows[r] or ()
        code = _str(row[0] if len(row) > 0 else "")
        if not code or not re.match(r"^\d{4,6}$", code):
            continue
        def g(i): return row[i] if len(row) > i else None
        out[code] = {
            "cbName":           _str(g(1)),
            "couponRate":       _num(g(3)),
            "convPrice":        _num(g(4)),
            "convEffDate":      _date_str(g(5)),
            "stockCode":        _str(g(6)),
            "stockName":        _str(g(7)),
            "convStart":        _date_str(g(8)),
            "convEnd":          _date_str(g(9)),
            "issueDate":        _date_str(g(10)),
            "listDate":         _date_str(g(11)),
            "maturityDate":     _date_str(g(12)),
            "maturityPrice":    _num(g(13)),
            "issueTotal":       _num(g(15)),
            "actualTotal":      _num(g(16)),
            "issuePrice":       _num(g(17)),
            "latestBalance":    _num(g(18)),
            "repayYears":       _num(g(19)),
            "issueConvPrice":   _num(g(20)),
            "underwriter":      _str(g(21)),
            "guarantee":        _str(g(39)),
            "nearestPutDate":   _date_str(g(36)),
            "nearestPutPrice":  _num(g(37)),
            "nearestPutYield":  _num(g(38)),
            "resetFormula":     _str(g(41)),
            "callDate":         _date_str(g(44)),
        }
    return out


def _parse_call_rights(rows: list[tuple]) -> list[dict]:
    """公司執行贖回權: header「公司代號」"""
    hdr = _find_header_row(rows, "公司代號")
    if hdr < 0:
        return []
    out: list[dict] = []
    for r in range(hdr + 1, len(rows)):
        row = rows[r] or ()
        code = _str(row[0] if len(row) > 0 else "")
        if not code:
            continue
        out.append({
            "stockCode":  code,
            "stockName":  _str(row[1] if len(row) > 1 else ""),
            "reportDate": _str(row[2] if len(row) > 2 else ""),
            "subject":    _str(row[3] if len(row) > 3 else ""),
            "asoExpiry":  _date_str(row[4] if len(row) > 4 else None),
        })
    return out


def _parse_conv_stop(rows: list[tuple]) -> list[dict]:
    """停止轉換資訊: header「債券代碼」"""
    hdr = _find_header_row(rows, "債券代碼")
    if hdr < 0:
        return []
    out: list[dict] = []
    for r in range(hdr + 1, len(rows)):
        row = rows[r] or ()
        code = _str(row[0] if len(row) > 0 else "")
        if not code or not re.match(r"^\d{4,6}$", code):
            continue
        out.append({
            "cbCode":    code,
            "cbName":    _str(row[1] if len(row) > 1 else ""),
            "startDate": _str(row[2] if len(row) > 2 else ""),
            "endDate":   _str(row[3] if len(row) > 3 else ""),
            "reason":    _str(row[4] if len(row) > 4 else ""),
        })
    return out


def _parse_price_adjust(rows: list[tuple]) -> list[dict]:
    """轉換價格調整: header「種類」"""
    hdr = _find_header_row(rows, "種類")
    if hdr < 0:
        return []
    out: list[dict] = []
    for r in range(hdr + 1, len(rows)):
        row = rows[r] or ()
        type_ = _str(row[0] if len(row) > 0 else "")
        if not type_:
            continue
        out.append({
            "type":       type_,
            "stockCode":  _str(row[1] if len(row) > 1 else ""),
            "stockName":  _str(row[2] if len(row) > 2 else ""),
            "reportDate": _str(row[3] if len(row) > 3 else ""),
            "subject":    _str(row[4] if len(row) > 4 else ""),
            "newPrice":   _num(row[5] if len(row) > 5 else None),
        })
    return out


def _parse_outstanding(rows: list[tuple]) -> dict[str, Any]:
    """CB流通在外餘額: header「證券代號」,資料從 header+2"""
    hdr = _find_header_row(rows, "證券代號")
    if hdr < 0:
        return {}
    out: dict[str, Any] = {}
    for r in range(hdr + 2, len(rows)):
        row = rows[r] or ()
        code = _str(row[0] if len(row) > 0 else "")
        if not code or not re.match(r"^\d{4,6}$", code):
            continue
        def g(i): return row[i] if len(row) > i else None
        out[code] = {
            "cbName":     _str(g(1)),
            "thisWeek":   _num(g(2)),
            "lastWeek":   _num(g(3)),
            "change":     _num(g(4)),
            "changeRate": _num(g(5)),
            "remainPct":  _num(g(6)),
        }
    return out


def _parse_daily_market(rows: list[tuple]) -> dict[str, Any]:
    """CB每日行情表: header「代碼」(只在 row[0])"""
    hdr = -1
    for r in range(min(5, len(rows))):
        row = rows[r] or ()
        if len(row) > 0 and _str(row[0]) == "代碼":
            hdr = r
            break
    if hdr < 0:
        return {}
    out: dict[str, Any] = {}
    for r in range(hdr + 1, len(rows)):
        row = rows[r] or ()
        code = _str(row[0] if len(row) > 0 else "")
        if not code or not re.match(r"^\d{4,6}$", code):
            continue
        def g(i): return row[i] if len(row) > i else None
        out[code] = {
            "cbName":      _str(g(1)),
            "industry":    _str(g(2)),
            "ytp":         _num(g(16)),
            "ytm":         _num(g(17)),
            "eps":         _num(g(18)),
            "vol120":      _num(g(19)),
            "vol240":      _num(g(20)),
            "tcri":        _num(g(25)),
            "issueTotal":  _num(g(26)),
            "outstanding": _num(g(27)),
            "remainPct":   _num(g(28)),
            "guarantee":   _str(g(29)),
            "business":    _str(g(30)),
        }
    return out


# ── Drive 取最新檔 ───────────────────────────────────────────────────
def _drive_service():
    """重用 lib.drive 的 service account credentials path。"""
    from lib import drive  # 避免 import 順序問題
    return drive._get_service()


def fetch_latest_xlsx(folder_id: str = DEFAULT_FOLDER_ID
                      ) -> tuple[str, bytes]:
    """從 Drive folder 撈檔名 '元大證選擇權YYYYMMDD.xlsx' 最大日期那份。

    回傳 (date_str, content_bytes)。找不到時 raise RuntimeError。
    """
    svc = _drive_service()
    files = svc.files().list(
        q=f"'{folder_id}' in parents and trashed = false",
        fields="files(id,name,modifiedTime)",
        pageSize=100,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])

    candidates: list[tuple[str, dict]] = []
    for f in files:
        m = FILE_NAME_RE.search(f.get("name", ""))
        if m:
            candidates.append((m.group(1), f))
    if not candidates:
        raise RuntimeError(
            f"folder {folder_id} 內找不到「元大證選擇權YYYYMMDD.xlsx」"
        )
    candidates.sort(key=lambda x: x[0], reverse=True)
    date_str, target = candidates[0]

    # 下載
    from googleapiclient.http import MediaIoBaseDownload
    req = svc.files().get_media(fileId=target["id"], supportsAllDrives=True)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return date_str, buf.getvalue()


# ── 主入口 ────────────────────────────────────────────────────────────
def parse_xlsx(content: bytes, report_date: str = "") -> dict[str, Any]:
    """解析整份 xlsx → yuantaReport dict。GAS 流程的 Python 等價物。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    out: dict[str, Any] = {
        "reportDate":     report_date,
        "quotes":         {},
        "basicInfo":      {},
        "callRights":     [],
        "conversionStop": [],
        "priceAdjust":    [],
        "outstanding":    {},
        "dailyMarket":    {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        try:
            if sheet_name == "報價單":
                out["quotes"] = _parse_quotes(rows)
            elif sheet_name == "基本資料檔":
                out["basicInfo"] = _parse_basic_info(rows)
            elif sheet_name == "公司執行贖回權":
                out["callRights"] = _parse_call_rights(rows)
            elif sheet_name.startswith("停止轉換資訊"):
                out["conversionStop"] = _parse_conv_stop(rows)
            elif sheet_name == "轉換價格調整":
                out["priceAdjust"] = _parse_price_adjust(rows)
            elif sheet_name.startswith("CB流通在外餘額"):
                out["outstanding"] = _parse_outstanding(rows)
            elif sheet_name.startswith("CB每日行情表"):
                out["dailyMarket"] = _parse_daily_market(rows)
        except Exception as e:  # noqa: BLE001
            print(f"⚠️ 解析 sheet '{sheet_name}' 失敗: {e}",
                  file=sys.stderr, flush=True)

    # 過濾 outstanding: 只留有出現在 quotes/basicInfo/dailyMarket 的 CB
    known: set[str] = set()
    known.update(out["quotes"].keys())
    known.update(out["basicInfo"].keys())
    known.update(out["dailyMarket"].keys())
    out["outstanding"] = {k: v for k, v in out["outstanding"].items() if k in known}

    return out


def fetch_and_parse(folder_id: str = DEFAULT_FOLDER_ID) -> dict[str, Any]:
    """一站式:Drive 取最新檔 → 解析 → 回 yuantaReport dict。"""
    date_str, blob = fetch_latest_xlsx(folder_id)
    return parse_xlsx(blob, report_date=date_str)


# ── Smoke test ────────────────────────────────────────────────────────
def _smoke() -> int:
    folder_id = os.environ.get("YUANTA_REPORT_FOLDER_ID", DEFAULT_FOLDER_ID)
    print(f"folder_id = {folder_id}")
    res = fetch_and_parse(folder_id)
    print(f"reportDate     = {res['reportDate']}")
    print(f"quotes         = {len(res['quotes']):,} CBs")
    print(f"basicInfo      = {len(res['basicInfo']):,} CBs")
    print(f"callRights     = {len(res['callRights']):,} entries")
    print(f"conversionStop = {len(res['conversionStop']):,} entries")
    print(f"priceAdjust    = {len(res['priceAdjust']):,} entries")
    print(f"outstanding    = {len(res['outstanding']):,} CBs")
    print(f"dailyMarket    = {len(res['dailyMarket']):,} CBs")

    # 統計 callDate / resetFormula 非空
    bi = res["basicInfo"]
    n_call = sum(1 for v in bi.values() if v.get("callDate"))
    n_reset = sum(1 for v in bi.values() if v.get("resetFormula"))
    print(f"\n[basicInfo] callDate 非空: {n_call} / {len(bi)}")
    print(f"[basicInfo] resetFormula 非空: {n_reset} / {len(bi)}")
    if n_call:
        print("有強制贖回的 CB:")
        for k, v in bi.items():
            if v.get("callDate"):
                print(f"  {k} {v['cbName']:>10s}  callDate={v['callDate']}")
    return 0


if __name__ == "__main__":
    sys.exit(_smoke())
